from dataclasses import dataclass, field
from pathlib import Path
import openpyxl
import pandas as pd

DEFAULT_WORKBOOK = Path(__file__).parent / "energy_model.xlsx"

@dataclass
class WorkbookData:
    saudi_data: dict
    nuclear_tech: pd.DataFrame
    scenarios: pd.DataFrame
    weights: dict
    carbon: dict
    forecast_seed: pd.DataFrame

def _sheet_to_df(ws) -> pd.DataFrame:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        non_empty = [v for v in row if v not in (None, "")]
        if len(non_empty) >= 2 and all(isinstance(v, str) for v in non_empty[:2]):
            header_idx = i
            break
    if header_idx is None:
        header_idx = 0
    header = [str(v) if v is not None else f"col{i}" for i, v in enumerate(rows[header_idx])]
    data_rows = [r for r in rows[header_idx + 1:] if any(v not in (None, "") for v in r)]
    df = pd.DataFrame(data_rows, columns=header)
    return df.dropna(axis=1, how="all")

def load_workbook(path: str | Path = DEFAULT_WORKBOOK) -> WorkbookData:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(
            f"Workbook not found at {path}. Place the reference .xlsx next to "
            "this script or pass an explicit --workbook path."
        )
    wb = openpyxl.load_workbook(path, data_only=True)
    
    ws = wb["Saudi_Data"]
    saudi_data = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        name, value = row[0], row[1]
        if isinstance(value, (int, float)):
            saudi_data[name] = value
        elif isinstance(value, str) and value.strip().endswith("%"):
            try:
                saudi_data[name] = float(value.strip().rstrip("%")) / 100
            except ValueError:
                pass

    nuclear_tech = _sheet_to_df(wb["Nuclear_Technologies"])
    scenarios = _sheet_to_df(wb["Scenario_Simulator"])
    for col in ["Demand Multiplier", "Water", "Renewables"]:
        if col in scenarios.columns:
            scenarios[col] = pd.to_numeric(scenarios[col], errors="coerce")

    ws = wb["Decision_Engine"]
    weights = {}
    for row in ws.iter_rows(min_row=4, max_row=8, values_only=True):
        if row and row[0] and isinstance(row[1], (int, float)):
            weights[row[0]] = row[1]

    ws = wb["Carbon_Emissions"]
    carbon = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row and row[0] and isinstance(row[1], (int, float)):
            carbon[row[0]] = row[1]

    forecast_seed = _sheet_to_df(wb["AI_Forecast"])

    return WorkbookData(
        saudi_data=saudi_data,
        nuclear_tech=nuclear_tech,
        scenarios=scenarios,
        weights=weights,
        carbon=carbon,
        forecast_seed=forecast_seed,
    )

if __name__ == "__main__":
    data = load_workbook()
    print("Saudi Data variables loaded:", len(data.saudi_data))
    for k, v in data.saudi_data.items():
        print(f"  {k}: {v}")
    print("\nScenarios:\n", data.scenarios)
    print("\nWeights:", data.weights)
    print("\nCarbon intensities:", data.carbon)
